"""Write XLSX workbooks from ExtractedDocument / StructuredRecord.

Default workbook structure (no template):
  Summary      one row per document (fields + metadata)
  Table_001    one sheet per detected table
  Pages        page-level text when no tables found
  Warnings     extraction warnings

When a template path is provided, rows are appended starting from the
configured start row — used for multi-document batch output.
"""

from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from pdf_to_markdown.structured.models import ExtractedDocument, StructuredRecord

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="1A2744")
_ALT_FILL = PatternFill(fill_type="solid", fgColor="F1F5F9")
_WARN_FILL = PatternFill(fill_type="solid", fgColor="FEF3C7")


def _style_header(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(wrap_text=False)
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _auto_width(ws, max_width: int = 60) -> None:
    for col in ws.columns:
        width = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 4, max_width)


def _append_table_sheet(wb: Workbook, table, sheet_name: str) -> None:
    ws = wb.create_sheet(sheet_name)
    cols = table.columns or [f"col_{i}" for i in range(len(table.rows[0]))] if table.rows else []
    if not cols:
        return
    ws.append(["page"] + cols)
    for i, row in enumerate(table.rows):
        ws.append([table.page_number or ""] + row)
        if i % 2 == 1:
            for cell in ws[ws.max_row]:
                cell.fill = _ALT_FILL
    _style_header(ws)
    _auto_width(ws)


def write_xlsx(
    doc: ExtractedDocument,
    record: StructuredRecord,
    out_path: Path,
) -> Path:
    """Write a self-contained workbook for one document."""
    wb = Workbook()
    wb.remove(wb.active)  # remove default Sheet

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary")
    meta_headers = [
        "source_file", "document_id", "document_type", "pages",
        "tables_found", "fields_found", "warnings", "extracted_at",
        "extraction_method",
    ]
    ws_sum.append(meta_headers)
    ws_sum.append([
        doc.source_file,
        doc.document_id,
        record.document_type,
        doc.page_count,
        len(doc.tables),
        len(record.fields),
        len(doc.warnings),
        doc.extracted_at.isoformat(),
        doc.extraction_method,
    ])
    # Append extracted fields as additional rows in the same sheet
    if record.fields:
        ws_sum.append([])
        ws_sum.append(["field_name", "value"])
        for k, v in record.fields.items():
            ws_sum.append([k, str(v)])
    _style_header(ws_sum)
    _auto_width(ws_sum)

    # ── One sheet per table ───────────────────────────────────────────────────
    for idx, table in enumerate(doc.tables, start=1):
        name = f"Table_{idx:03d}"
        if table.rows:
            _append_table_sheet(wb, table, name)

    # ── Pages sheet (fallback when no tables) ─────────────────────────────────
    if not doc.tables:
        ws_pages = wb.create_sheet("Pages")
        ws_pages.append(["page_number", "text", "extraction_method"])
        for p in doc.pages:
            if p.text.strip():
                ws_pages.append([p.page_number, p.text, p.extraction_method])
        _style_header(ws_pages)
        _auto_width(ws_pages)

    # ── Warnings sheet (always present) ──────────────────────────────────────
    ws_warn = wb.create_sheet("Warnings")
    ws_warn.append(["code", "message", "page_number"])
    for w in doc.warnings:
        row = ws_warn.max_row + 1
        ws_warn.append([w.code, w.message, w.page_number or ""])
        for cell in ws_warn[row]:
            cell.fill = _WARN_FILL
    _style_header(ws_warn)
    _auto_width(ws_warn)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return out_path


# ── Batch / template writer ───────────────────────────────────────────────────

def write_batch_xlsx(
    results: list[tuple[ExtractedDocument, StructuredRecord]],
    out_path: Path,
    template_path: Optional[Path] = None,
) -> Path:
    """Append all records into one workbook (single writer, never concurrent).

    If template_path is provided, load it and append rows starting from the
    first empty row of each sheet. Otherwise generate a default workbook.
    """
    if template_path and template_path.exists():
        wb = load_workbook(str(template_path))
    else:
        wb = Workbook()
        wb.remove(wb.active)
        # Create standard sheets
        for name in ("Summary", "LineItems", "Warnings", "RawPreview"):
            wb.create_sheet(name)

    ws_summary = wb["Summary"] if "Summary" in wb.sheetnames else wb.create_sheet("Summary")
    ws_items = wb["LineItems"] if "LineItems" in wb.sheetnames else wb.create_sheet("LineItems")
    ws_warnings = wb["Warnings"] if "Warnings" in wb.sheetnames else wb.create_sheet("Warnings")

    # Write headers if sheets are empty
    if ws_summary.max_row <= 1 and not any(ws_summary[1]):
        ws_summary.append([
            "source_file", "document_type", "pages", "tables_found",
            "fields_found", "warnings", "extraction_method", "extracted_at",
        ])
        _style_header(ws_summary)

    if ws_items.max_row <= 1 and not any(ws_items[1]):
        ws_items.append(["source_file", "document_id", "table_id", "row_index"] + ["col_0"])
        _style_header(ws_items)

    if ws_warnings.max_row <= 1 and not any(ws_warnings[1]):
        ws_warnings.append(["source_file", "code", "message", "page_number"])
        _style_header(ws_warnings)

    for doc, record in results:
        ws_summary.append([
            doc.source_file,
            record.document_type,
            doc.page_count,
            len(doc.tables),
            len(record.fields),
            len(doc.warnings),
            doc.extraction_method,
            doc.extracted_at.isoformat(),
        ])

        for table in doc.tables:
            cols = table.columns or [f"col_{i}" for i in range(len(table.rows[0]))] if table.rows else []
            for row_idx, row in enumerate(table.rows):
                ws_items.append([doc.source_file, doc.document_id, table.table_id, row_idx] + row)

        for w in doc.warnings:
            ws_warnings.append([doc.source_file, w.code, w.message, w.page_number or ""])

    _auto_width(ws_summary)
    _auto_width(ws_items)
    _auto_width(ws_warnings)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return out_path
