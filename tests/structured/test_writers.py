"""Tests for CSV and XLSX writers using synthetic ExtractedDocument."""

import csv
from pathlib import Path

import openpyxl
import pytest

from pdf_to_markdown.structured.csv_writer import write_csv_outputs
from pdf_to_markdown.structured.models import (
    ExtractedDocument,
    ExtractedPage,
    ExtractedTable,
    ExtractionWarning,
    StructuredRecord,
    TextBlock,
)
from pdf_to_markdown.structured.xlsx_writer import write_batch_xlsx, write_xlsx


def _make_doc_with_table() -> tuple[ExtractedDocument, StructuredRecord]:
    table = ExtractedTable(
        page_number=1,
        columns=["Description", "Amount"],
        rows=[["Freight", "1200.00"], ["Fuel surcharge", "120.00"]],
        source="pymupdf",
        confidence=0.9,
    )
    doc = ExtractedDocument(
        source_file="invoice.pdf",
        page_count=2,
        pages=[ExtractedPage(page_number=1, text="Invoice Number: INV-001", tables=[table])],
        tables=[table],
        text_blocks=[TextBlock(page_number=1, text="Invoice Number: INV-001")],
        warnings=[],
    )
    record = StructuredRecord(
        document_id=doc.document_id,
        source_file="invoice.pdf",
        fields={"Invoice Number": "INV-001", "Total Amount": "1320.00"},
        line_items=[{"Description": "Freight", "Amount": "1200.00"}],
    )
    return doc, record


def _make_doc_no_table() -> tuple[ExtractedDocument, StructuredRecord]:
    doc = ExtractedDocument(
        source_file="text_only.pdf",
        page_count=1,
        pages=[ExtractedPage(page_number=1, text="Just some paragraph text.")],
        tables=[],
        text_blocks=[TextBlock(page_number=1, text="Just some paragraph text.")],
        warnings=[ExtractionWarning(code="NO_TABLES", message="No tables detected")],
    )
    record = StructuredRecord(
        document_id=doc.document_id, source_file="text_only.pdf"
    )
    return doc, record


# ── CSV tests ────────────────────────────────────────────────────────────────

def test_csv_table_written(tmp_path):
    doc, record = _make_doc_with_table()
    written = write_csv_outputs(doc, record, tmp_path, "invoice")
    table_csv = tmp_path / "invoice__table_001.csv"
    assert table_csv.exists()
    rows = list(csv.reader(table_csv.open()))
    assert rows[0] == ["page", "Description", "Amount"]
    assert rows[1][1] == "Freight"


def test_csv_fields_written(tmp_path):
    doc, record = _make_doc_with_table()
    write_csv_outputs(doc, record, tmp_path, "invoice")
    fields_csv = tmp_path / "invoice__fields.csv"
    assert fields_csv.exists()
    rows = list(csv.reader(fields_csv.open()))
    field_names = [r[0] for r in rows[1:]]
    assert "Invoice Number" in field_names


def test_csv_pages_fallback_when_no_tables(tmp_path):
    doc, record = _make_doc_no_table()
    write_csv_outputs(doc, record, tmp_path, "text_only")
    assert (tmp_path / "text_only__pages.csv").exists()
    assert not (tmp_path / "text_only__table_001.csv").exists()


def test_csv_warnings_always_written(tmp_path):
    doc, record = _make_doc_no_table()
    write_csv_outputs(doc, record, tmp_path, "text_only")
    warn_csv = tmp_path / "text_only__warnings.csv"
    assert warn_csv.exists()
    rows = list(csv.reader(warn_csv.open()))
    assert rows[1][0] == "NO_TABLES"


# ── XLSX tests ────────────────────────────────────────────────────────────────

def test_xlsx_single_doc(tmp_path):
    doc, record = _make_doc_with_table()
    out = tmp_path / "out.xlsx"
    write_xlsx(doc, record, out)
    assert out.exists()
    wb = openpyxl.load_workbook(str(out))
    assert "Summary" in wb.sheetnames
    assert "Table_001" in wb.sheetnames
    assert "Warnings" in wb.sheetnames


def test_xlsx_table_sheet_has_data(tmp_path):
    doc, record = _make_doc_with_table()
    out = tmp_path / "out.xlsx"
    write_xlsx(doc, record, out)
    wb = openpyxl.load_workbook(str(out))
    ws = wb["Table_001"]
    # Row 1 = header, row 2+ = data
    assert ws.cell(1, 1).value == "page"
    assert ws.cell(2, 2).value == "Freight"


def test_xlsx_pages_sheet_when_no_tables(tmp_path):
    doc, record = _make_doc_no_table()
    out = tmp_path / "out.xlsx"
    write_xlsx(doc, record, out)
    wb = openpyxl.load_workbook(str(out))
    assert "Pages" in wb.sheetnames
    assert "Table_001" not in wb.sheetnames


def test_xlsx_batch_two_docs(tmp_path):
    pair1 = _make_doc_with_table()
    pair2 = _make_doc_no_table()
    out = tmp_path / "batch.xlsx"
    write_batch_xlsx([pair1, pair2], out)
    wb = openpyxl.load_workbook(str(out))
    ws = wb["Summary"]
    # Header + 2 data rows
    assert ws.max_row >= 3
